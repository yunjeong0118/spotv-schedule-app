"""
SPOTV 주간 편성표(.xlsx) 파싱 → 업로드파일 데이터 변환 핵심 로직.

이 모듈은 순수 규칙 기반(결정적) 로직이며 AI 호출이 필요 없다.
검증된 규칙:
  1) 세로축(시간)은 B열의 시 라벨(8행 간격, 03~24~01~02) + C열의 30분 마커로 결정된다.
  2) 24시를 넘는 라벨(24,01,02)은 다음 날 새벽으로, 날짜 +1일 처리한다.
  3) 프로그램 블록이 새 항목인지 / 직전 항목의 연속(대진정보 등)인지는
     "셀 테두리(위/아래 border)"로 판정한다. 둘 다 없으면 연속(같은 편성 구간).
  4) 각 날짜 열의 첫 블록 top border가 없으면, 실제 방송 시작은 그리드상의 03:00이 아니라
     "전날(왼쪽) 열의 마지막 실제 편성이 끝나고 빈칸이 시작되는 시각"부터다.
     (전날 열이 없거나 전날 열이 끝까지 채워져 있으면 30분 앞당기는 것으로 대체 추정한다.)
  5) 셀 텍스트에 "생중계"/"생방송"이라는 단어가 있으면(앞에 HH:MM이 붙어 있든 없든)
     - 그 단어(및 붙어있는 시각)는 프로그램명(str7)에서 반드시 제외한다.
     - 방송구분(str8)은 LIVE로 확정한다.
     - 시각이 명시되어 있으면(HH:MM) 그 시각을 그대로 채택하고, 없으면 그리드 위치로 추정한다.
  6) 방송구분(본방송/재방송/LIVE)은 기본적으로 셀 배경색(theme)으로 판정한다.
     현장생중계·수신생중계=LIVE, 녹화중계·본방송=본방송, 배경없음=재방송.
     단 5)의 생중계/생방송 텍스트가 있으면 색상과 무관하게 LIVE로 확정한다.
  7) 자막(자)/해설(해)/수어(수) 표기는 셀 값이 아니라 도형(텍스트박스)이며,
     xlsx 내부의 xl/drawings/drawing*.xml 을 파싱해서 좌표로 매칭해야 한다.
  8) 연령고지 원문자(⑮ 등)는 프로그램명에서 제거한다.
"""

import io
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date, timedelta, datetime

import openpyxl

DAY_COLS_DEFAULT = [4, 6, 8, 10, 12, 14, 16]  # D,F,H,J,L,N,P

# HH:MM은 있어도 없어도 되고, "생중계"/"생방송" 뒤에 이어지는 텍스트(있으면)를 4번째 그룹으로 잡는다.
LIVE_MARKER_RE = re.compile(r'^\s*(?:(\d{1,2}):(\d{2})\s*)?(생중계|생방송)\s*(.*)$', re.S)

NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
}


def clean(text):
    if text is None:
        return ''
    t = text.replace('\n', ' ')
    t = re.sub(r'\[HD\]', '', t)
    # 연령고지(⑮ 등 원문자 등급 표기) 제거
    t = re.sub(r'[\u2460-\u2473\u24EA\u3251-\u325F\u32B1-\u32BF]', '', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def extract_live_marker(ct):
    """정리된 텍스트에서 '생중계'/'생방송' 표기를 분리한다.
    반환: (is_live, explicit_hour|None, explicit_minute|None, 나머지_텍스트)
    나머지_텍스트에는 '생중계'/'생방송'과 그 앞의 HH:MM이 제거되어 있다.
    """
    m = LIVE_MARKER_RE.match(ct)
    if not m:
        return False, None, None, ct
    h = int(m.group(1)) if m.group(1) is not None else None
    mi = int(m.group(2)) if m.group(2) is not None else None
    trailing = m.group(4).strip()
    return True, h, mi, trailing


def _fillcat(cell):
    fill = cell.fill
    if fill.patternType is None:
        return '재방송'
    fg = fill.fgColor
    if fg.type == 'theme':
        theme = fg.theme
        # 현장생중계, 수신생중계 -> LIVE
        if theme in (3, 5):
            return 'LIVE'
        # 녹화중계, 본방송 -> 본방송
        if theme in (2, 7):
            return '본방송'
    return 'UNKNOWN'


def _has_border(cell_border_side):
    return cell_border_side is not None and cell_border_side.style is not None


def find_start_date(ws, day_cols):
    """헤더 행(보통 6행)에서 각 날짜 열의 datetime 값을 읽어 date 딕셔너리로 반환."""
    day_dates = {}
    for col in day_cols:
        for row in range(1, 10):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, datetime):
                day_dates[col] = v.date()
                break
    return day_dates


def _build_row_to_time(ws):
    b_merges = sorted(
        [(mc.min_row, mc.max_row, ws.cell(row=mc.min_row, column=2).value)
         for mc in ws.merged_cells.ranges if mc.min_col == 2],
        key=lambda x: x[0]
    )
    hour_bounds = []
    past_midnight = False
    for i, (r1, r2, label) in enumerate(b_merges):
        try:
            hour = int(label)
        except (TypeError, ValueError):
            continue
        if hour == 24:
            past_midnight = True
        end = b_merges[i + 1][0] if i + 1 < len(b_merges) else r1 + 8
        h = hour % 24
        hour_bounds.append((r1, end, h, 1 if past_midnight else 0))

    def row_to_time(row):
        for start, end, hour, extra_days in hour_bounds:
            if start <= row < end:
                mid = start + 4
                minute = 0 if row < mid else 30
                return hour, minute, extra_days
        return None, None, None

    return row_to_time


def _shift_back(h, m, extra, minutes=30):
    total = h * 60 + m - minutes
    day_extra, total = divmod(total, 24 * 60)
    return total // 60, total % 60, extra + day_extra


def _get_column_blocks(ws, col, header_row):
    blocks = []
    for mc in ws.merged_cells.ranges:
        if mc.min_col == col:
            val = ws.cell(row=mc.min_row, column=col).value
            blocks.append((mc.min_row, mc.max_row, val))
    blocks.sort(key=lambda x: x[0])
    return [b for b in blocks if b[0] > header_row]


def parse_schedule_grid(ws, day_cols=None, first_block_shift_minutes=30):
    """워크시트 하나를 파싱해서 항목 리스트를 반환한다.
    각 항목: {date, hour, minute, title, cat, rows:[(r1,r2),...], col, needs_review}

    주의(중요한 한계): 어떤 날짜 열의 '첫 블록'에 위쪽 테두리가 없으면, 그 프로그램은
    그리드가 보여주는 시각(예: 03:00)보다 실제로는 더 일찍 시작된 것이다. 그런데 정확히
    몇 분/몇 시간 앞서는지는 파일 구조(테두리·색상·인접 열의 빈칸 등) 만으로는 안전하게
    특정할 수 없다는 것이 여러 실제 파일 대조로 확인됐다 — 같은 신호가 파일에 따라
    30분 앞일 수도, 1시간 앞일 수도 있다. 그래서 이 함수는 기본값(first_block_shift_minutes,
    기본 30분)으로 추정하되, 해당 항목에 needs_review=True 를 표시해서 사람이 실제 편성표를
    보고 확인하도록 한다. 화면(app.py)에서는 이 항목들을 목록으로 보여준다.
    """
    if day_cols is None:
        day_cols = DAY_COLS_DEFAULT
    day_dates = find_start_date(ws, day_cols)
    row_to_time = _build_row_to_time(ws)

    all_entries = []

    for idx, col in enumerate(day_cols):
        basedate = day_dates.get(col)
        if basedate is None:
            continue

        header_row = max([r for r in range(1, 10)
                           if isinstance(ws.cell(row=r, column=col).value, datetime)] or [6])
        blocks = _get_column_blocks(ws, col, header_row)

        entries = []
        pending = None
        first_block_seen = False
        prev_had_content = False
        prev_last_row = None

        def flush():
            nonlocal pending
            if pending and pending['title_parts']:
                title = ' '.join(p for p in pending['title_parts'] if p).strip()
                title = re.sub(r'\s+', ' ', title)
                cat = 'LIVE' if pending.get('forced_live') else pending['cat']
                entries.append({
                    'date': basedate + timedelta(days=pending['extra_days']),
                    'hour': pending['h'], 'minute': pending['m'],
                    'title': title, 'cat': cat,
                    'rows': pending['rows'],
                    'needs_review': pending.get('needs_review', False),
                })
            pending = None

        for (r1, r2, val) in blocks:
            text = val if isinstance(val, str) else None
            if text is None or clean(text) == '':
                flush()
                prev_had_content = False
                prev_last_row = None
                continue
            ct = clean(text)
            is_live, ex_h, ex_m, ct2 = extract_live_marker(ct)
            is_first_block = not first_block_seen
            first_block_seen = True

            if pending is None:
                start_fresh = True
            elif not prev_had_content:
                start_fresh = True
            else:
                top = ws.cell(row=r1, column=col).border.top
                bottom = ws.cell(row=prev_last_row, column=col).border.bottom
                start_fresh = _has_border(top) or _has_border(bottom)

            if start_fresh:
                flush()
                cat = _fillcat(ws.cell(row=r1, column=col))
                needs_review = False
                if ex_h is not None:
                    h, mnt = ex_h, ex_m
                    _, _, extra = row_to_time(r1)
                else:
                    h, mnt, extra = row_to_time(r1)
                    if is_first_block:
                        top = ws.cell(row=r1, column=col).border.top
                        if not _has_border(top):
                            h, mnt, extra = _shift_back(h, mnt, extra, first_block_shift_minutes)
                            needs_review = True
                pending = {'title_parts': [ct2] if ct2 else [], 'cat': cat,
                           'forced_live': is_live, 'needs_review': needs_review,
                           'h': h, 'm': mnt, 'extra_days': extra, 'rows': [(r1, r2)]}
            else:
                if ct2:
                    pending['title_parts'].append(ct2)
                if is_live:
                    pending['forced_live'] = True
                pending['rows'].append((r1, r2))

            prev_had_content = True
            prev_last_row = r2

        flush()
        for e in entries:
            e['col'] = col
        all_entries.extend(entries)

    all_entries.sort(key=lambda e: (e['date'], e['hour'], e['minute']))
    return all_entries


# ---------------------------------------------------------------------------
# 자막(자)/해설(해)/수어(수) 도형 태그 추출 및 매칭
# ---------------------------------------------------------------------------

def extract_shape_tags(file_bytes):
    """xlsx 파일(bytes)에서 도형 텍스트박스들의 (row, col, text)를 모두 추출한다.
    row/col은 1-indexed (openpyxl과 동일 좌표계).
    """
    shapes = []
    with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
        drawing_names = [n for n in z.namelist()
                          if n.startswith('xl/drawings/drawing') and n.endswith('.xml')]
        for name in drawing_names:
            with z.open(name) as f:
                tree = ET.parse(f)
            root = tree.getroot()
            for tag in ['twoCellAnchor', 'oneCellAnchor']:
                for anchor in root.findall(f'xdr:{tag}', NS):
                    frm = anchor.find('xdr:from', NS)
                    if frm is None:
                        continue
                    col_el = frm.find('xdr:col', NS)
                    row_el = frm.find('xdr:row', NS)
                    if col_el is None or row_el is None:
                        continue
                    col = int(col_el.text) + 1
                    row = int(row_el.text) + 1
                    sp = anchor.find('xdr:sp', NS)
                    if sp is None:
                        continue
                    texts = [t.text for t in sp.findall('.//a:t', NS) if t.text]
                    text = ''.join(texts).strip()
                    if text:
                        shapes.append((row, col, text))
    return shapes


def match_flags_to_entries(entries, shapes, day_cols=None, legend_rows=(5,)):
    """도형 태그를 항목에 매칭해서 각 entry['flags']에 {'자','해','수'} 세트를 채운다."""
    if day_cols is None:
        day_cols = DAY_COLS_DEFAULT

    def assign_daycol(tagcol):
        for dc in day_cols:
            if tagcol in (dc - 1, dc):
                return dc
        return None

    # 범례 영역(상단 행) 근처의 태그는 제외
    tagmap = defaultdict(set)
    for row, col, text in shapes:
        if text not in ('자', '해', '수'):
            continue
        if row in legend_rows:
            continue
        dc = assign_daycol(col)
        if dc is None:
            continue
        tagmap[(row, dc)].add(text)

    for e in entries:
        e['flags'] = set()
    by_col = defaultdict(list)
    for e in entries:
        by_col[e['col']].append(e)

    remaining = dict(tagmap)

    def try_pass(predicate):
        nonlocal remaining
        still = {}
        for (row, dc), tags in remaining.items():
            match = None
            for e in by_col[dc]:
                for (r1, r2) in e['rows']:
                    if predicate(row, r1, r2):
                        match = e
                        break
                if match:
                    break
            if match:
                match['flags'].update(tags)
            else:
                still[(row, dc)] = tags
        remaining = still

    try_pass(lambda row, r1, r2: r1 == row)
    try_pass(lambda row, r1, r2: r1 == row + 1)
    try_pass(lambda row, r1, r2: r1 <= row <= r2)
    return entries, remaining  # remaining = 끝까지 매칭 안 된 태그(디버그용)


def parse_full_schedule(file_bytes, day_cols=None, first_block_shift_minutes=30):
    """편성표 xlsx(bytes) 하나를 완전히 파싱: 항목 + 자/해/수 플래그까지 포함.
    첫 블록(테두리 없음) 보정폭은 기본 30분이며, 파일마다 실제 사실이 다를 수 있으므로
    해당 항목은 needs_review=True로 표시되어 화면에서 별도로 안내된다.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    entries = parse_schedule_grid(ws, day_cols=day_cols, first_block_shift_minutes=first_block_shift_minutes)
    shapes = extract_shape_tags(file_bytes)
    entries, unmatched = match_flags_to_entries(entries, shapes, day_cols=day_cols)
    return entries, unmatched


# ---------------------------------------------------------------------------
# 업로드파일 생성 / 변경사항 비교
# ---------------------------------------------------------------------------

def entry_key(e):
    return (e['date'].isoformat(), e['hour'], e['minute'])


def entry_dt(e):
    return datetime.combine(e['date'], datetime.min.time()) + timedelta(hours=e['hour'], minutes=e['minute'])


def entry_iyk(e):
    fl = e.get('flags', set())
    return ('Y' if '자' in fl else None,
            'Y' if '해' in fl else None,
            'Y' if '수' in fl else None)


def diff_entries(old_entries, new_entries):
    """(date,hour,minute) 키 기준으로 old/new를 비교해 변경 리스트를 만든다."""
    old_by_key = {entry_key(e): e for e in old_entries}
    new_by_key = {entry_key(e): e for e in new_entries}
    all_keys = sorted(set(old_by_key) | set(new_by_key))

    changes = []
    for k in all_keys:
        o = old_by_key.get(k)
        n = new_by_key.get(k)
        if o is None:
            changes.append({'kind': 'ADDED', 'key': k, 'old': None, 'new': n})
        elif n is None:
            changes.append({'kind': 'REMOVED', 'key': k, 'old': o, 'new': None})
        else:
            if (o['title'] != n['title'] or o['cat'] != n['cat']
                    or entry_iyk(o) != entry_iyk(n)):
                changes.append({'kind': 'CHANGED', 'key': k, 'old': o, 'new': n})
    return changes


WEEKDAY_KO = ['월', '화', '수', '목', '금', '토', '일']


def combine_moved(changes):
    """REMOVED와 ADDED 중 제목이 같은 것끼리는 '이동(MOVED)'으로 합친다."""
    removed = [c for c in changes if c['kind'] == 'REMOVED']
    added = [c for c in changes if c['kind'] == 'ADDED']
    used_added = set()
    combined = [c for c in changes if c['kind'] == 'CHANGED']

    for r in removed:
        match = None
        for i, a in enumerate(added):
            if i in used_added:
                continue
            if a['new']['title'] == r['old']['title']:
                match = (i, a)
                break
        if match:
            i, a = match
            used_added.add(i)
            combined.append({'kind': 'MOVED', 'key': r['key'], 'old': r['old'], 'new': a['new'],
                              'new_key': a['key']})
        else:
            combined.append(r)
    for i, a in enumerate(added):
        if i not in used_added:
            combined.append(a)
    return combined


def build_change_ranges(changes, new_entries_sorted, old_entries_sorted=None):
    """변경된 각 항목에 대해 "시작~종료" 구간을 계산하고 날짜별로 묶는다.
    종료 시각은 같은(new) 정렬 리스트에서 해당 항목 바로 다음 항목의 시작 시각으로 잡는다.
    REMOVED+ADDED로 제목이 같은 항목은 combine_moved()를 통해 MOVED로 합쳐서 하나의 구간으로 표시한다.
    """
    changes = combine_moved(changes)

    dt_list = [entry_dt(e) for e in new_entries_sorted]
    key_to_idx = {entry_key(e): i for i, e in enumerate(new_entries_sorted)}

    old_dt_list = [entry_dt(e) for e in old_entries_sorted] if old_entries_sorted else None
    old_key_to_idx = {entry_key(e): i for i, e in enumerate(old_entries_sorted)} if old_entries_sorted else None

    by_date = defaultdict(list)
    notes = []

    for ch in changes:
        if ch['kind'] == 'REMOVED':
            o = ch['old']
            notes.append(f"{o['date'].strftime('%m/%d')} {o['hour']:02d}:{o['minute']:02d} "
                          f"'{o['title']}' 편성 삭제(다른 시간대로 이동했을 수 있음)")
            continue

        if ch['kind'] == 'MOVED':
            # 시작: 옛 파일에서 old 항목 시작 / 종료: 새 파일에서 new 항목 종료
            start = entry_dt(ch['old'])
            if old_dt_list is not None and ch['key'] in old_key_to_idx:
                oi = old_key_to_idx[ch['key']]
                # 종료는 새 위치 기준으로 잡는다(새 편성표가 최신 기준이므로)
            idx = key_to_idx.get(ch['new_key'])
            end = dt_list[idx + 1] if (idx is not None and idx + 1 < len(dt_list)) else None
            by_date[start.date()].append((start, end, ch))
            continue

        idx = key_to_idx.get(ch['key'])
        if idx is None:
            continue
        start = dt_list[idx]
        end = dt_list[idx + 1] if idx + 1 < len(dt_list) else None
        by_date[start.date()].append((start, end, ch))

    lines = []
    detail_rows = []
    for d in sorted(by_date):
        ranges = sorted(by_date[d], key=lambda x: x[0])
        parts = []
        for start, end, ch in ranges:
            s = start.strftime('%H:%M')
            e_ = end.strftime('%H:%M') if end else '?'
            if end and end.date() != start.date():
                e_ += '(익일)'
            parts.append(f"{s} ~ {e_}")
            n = ch['new']
            o = ch['old']
            if ch['kind'] == 'ADDED':
                desc = f"신규 편성: {n['title']} ({n['cat']})"
            elif ch['kind'] == 'MOVED':
                desc = (f"'{n['title']}' 편성 시간 이동: "
                        f"{o['hour']:02d}:{o['minute']:02d} → {n['hour']:02d}:{n['minute']:02d}")
            else:
                desc = f"{o['title']} ({o['cat']}) → {n['title']} ({n['cat']})"
            detail_rows.append((d, f"{s}~{e_}", desc))
        wd = WEEKDAY_KO[d.weekday()]
        lines.append(f"- {d.month}/{d.day}({wd}) " + " / ".join(parts))

    report_text = "#변경구간\n" + "\n".join(lines)
    if notes:
        report_text += "\n\n[참고]\n" + "\n".join(f"- {n}" for n in notes)
    return report_text, detail_rows


def write_upload_excel(new_entries, old_entries=None, sheet_title='SPOTV_업로드파일'):
    """openpyxl Workbook을 반환한다. old_entries가 주어지면 변경된 셀만 노란색으로 표시.
    needs_review=True인 항목(첫 블록 테두리 없음으로 시각을 추정한 경우)은 주황색으로 표시한다.
    """
    from openpyxl.styles import Font, PatternFill

    new_sorted = sorted(new_entries, key=entry_dt)
    old_by_key = {entry_key(e): e for e in old_entries} if old_entries else {}

    YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    ORANGE = PatternFill(start_color='FFD8A8', end_color='FFD8A8', fill_type='solid')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = [f'str{i}' for i in range(1, 12)]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(name='Arial')

    for i, e in enumerate(new_sorted, start=1):
        row = i + 1
        flags = e.get('flags', set())
        i_val = 'Y' if '자' in flags else None
        j_val = 'Y' if '해' in flags else None
        k_val = 'Y' if '수' in flags else None
        values = [str(i), e['date'].year, f"{e['date'].month:02d}", f"{e['date'].day:02d}",
                  e['hour'], e['minute'], e['title'], e['cat'], i_val, j_val, k_val]

        changed_cols = set()
        if old_entries is not None:
            o = old_by_key.get(entry_key(e))
            if o is None:
                changed_cols = set(range(1, 12))
            else:
                oi, oj, ok = entry_iyk(o)
                if o['title'] != e['title']:
                    changed_cols.add(7)
                if o['cat'] != e['cat']:
                    changed_cols.add(8)
                if oi != i_val:
                    changed_cols.add(9)
                if oj != j_val:
                    changed_cols.add(10)
                if ok != k_val:
                    changed_cols.add(11)

        review = e.get('needs_review', False)

        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = Font(name='Arial')
            if c in (5, 6):
                cell.number_format = '00'
            if c in changed_cols:
                cell.fill = YELLOW
            elif review and c in (5, 6):
                cell.fill = ORANGE

    widths = {1: 6, 2: 6, 3: 5, 4: 5, 5: 5, 6: 5, 7: 55, 8: 9, 9: 6, 10: 6, 11: 6}
    for c, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    return wb
